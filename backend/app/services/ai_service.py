import io
import json
import math
import re
from typing import Any
from uuid import UUID

import httpx
from fastapi import HTTPException, status
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import get_settings
from ..models.entities import Collection, CollectionItem, Resource, ResourceChunk, User


class AIService:
    """Low-risk Gemini-backed AI and RAG service for this project."""

    def _gemini_headers(self) -> dict[str, str]:
        settings = get_settings()
        if not settings.gemini_api_key:
            raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, "Gemini API key is not configured")
        return {"Content-Type": "application/json", "x-goog-api-key": settings.gemini_api_key}

    def _gemini_base_url(self) -> str:
        base_url = get_settings().gemini_base_url.rstrip("/")
        return base_url if base_url.endswith("/v1beta") else f"{base_url}/v1beta"

    def _gemini_generate_url(self) -> str:
        settings = get_settings()
        return f"{self._gemini_base_url()}/models/{settings.gemini_model}:generateContent"

    def _gemini_embed_url(self) -> str:
        settings = get_settings()
        return f"{self._gemini_base_url()}/models/{settings.gemini_embedding_model}:embedContent"

    async def _call_gemini(self, prompt: str, *, max_output_tokens: int = 1200, temperature: float = 0.2) -> str:
        headers = self._gemini_headers()
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": temperature, "maxOutputTokens": max_output_tokens},
        }

        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(30.0, connect=10.0)) as client:
                response = await client.post(self._gemini_generate_url(), headers=headers, json=payload)
        except httpx.TimeoutException as exc:
            raise HTTPException(status.HTTP_504_GATEWAY_TIMEOUT, "Gemini request timed out") from exc
        except httpx.RequestError as exc:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Gemini request failed: {exc}") from exc

        if response.status_code == 429:
            raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Gemini API rate limit reached")
        if response.status_code >= 400:
            detail = "Gemini provider request failed"
            try:
                error_payload = response.json()
                detail = error_payload.get("error", {}).get("message") or json.dumps(error_payload)
            except ValueError:
                detail = response.text
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Gemini provider error: {detail}")

        try:
            data = response.json()
        except ValueError as exc:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, "Gemini returned an empty or invalid response") from exc

        candidates = data.get("candidates") or []
        parts: list[str] = []
        for candidate in candidates:
            for part in candidate.get("content", {}).get("parts", []) or []:
                if isinstance(part, dict) and part.get("text"):
                    parts.append(part["text"])

        if not parts:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, "Gemini returned an empty response")
        return "\n".join(parts).strip()

    async def _embed_text(self, text: str) -> list[float]:
        clean_text = (text or "").strip()
        if not clean_text:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Cannot create an embedding for empty text")
        headers = self._gemini_headers()
        payload = {"content": {"parts": [{"text": clean_text}]}}

        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(30.0, connect=10.0)) as client:
                response = await client.post(self._gemini_embed_url(), headers=headers, json=payload)
        except httpx.TimeoutException as exc:
            raise HTTPException(status.HTTP_504_GATEWAY_TIMEOUT, "Embedding request timed out") from exc
        except httpx.RequestError as exc:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Embedding request failed: {exc}") from exc

        if response.status_code == 429:
            raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Gemini embedding rate limit reached")
        if response.status_code >= 400:
            detail = "Embedding provider request failed"
            try:
                error_payload = response.json()
                detail = error_payload.get("error", {}).get("message") or json.dumps(error_payload)
            except ValueError:
                detail = response.text
            detail = detail or f"HTTP {response.status_code}"
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Embedding provider error: {detail}")

        try:
            data = response.json()
        except ValueError as exc:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, "Embedding provider returned an empty or invalid response") from exc

        embedding = data.get("embedding")
        if embedding is None and isinstance(data.get("embeddings"), list) and data["embeddings"]:
            embedding = data["embeddings"][0]
        if isinstance(embedding, dict):
            values = embedding.get("values") or embedding.get("vector")
        elif isinstance(embedding, list):
            values = embedding
        else:
            values = None

        if not values:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, "Embedding provider returned no vector values")

        try:
            return [float(value) for value in values]
        except (TypeError, ValueError) as exc:
            raise HTTPException(status.HTTP_502_BAD_GATEWAY, "Embedding provider returned invalid vector data") from exc

    async def ensure_resource_access(self, db: AsyncSession, user: User, resource_id: UUID) -> Resource:
        resource = await db.get(Resource, resource_id)
        if resource is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Resource not found")

        if user.role.value == "admin":
            return resource

        allowed = await db.scalar(
            select(CollectionItem.resource_id)
            .join(Collection, Collection.id == CollectionItem.collection_id)
            .where(CollectionItem.resource_id == resource_id, Collection.user_id == user.id)
            .limit(1)
        )
        if allowed is None:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Resource access denied")
        return resource

    def _extract_text_from_bytes(self, filename: str | None, content: bytes) -> str:
        lower_name = (filename or "").lower()

        if lower_name.endswith(".txt"):
            return content.decode("utf-8", errors="ignore")

        if lower_name.endswith(".pdf"):
            try:
                from pypdf import PdfReader
            except ImportError as exc:  # pragma: no cover
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "PDF support requires pypdf to be installed") from exc
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for page in reader.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    pages.append(page_text)
            return "\n".join(pages)

        if lower_name.endswith(".docx"):
            try:
                from docx import Document
            except ImportError as exc:  # pragma: no cover
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "DOCX support requires python-docx to be installed") from exc
            document = Document(io.BytesIO(content))
            paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text and paragraph.text.strip()]
            return "\n".join(paragraphs)

        if lower_name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover
                raise HTTPException(status.HTTP_400_BAD_REQUEST, "XLSX support requires openpyxl to be installed") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        rows.append(" | ".join(values))
            return "\n".join(rows)

        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unsupported file type")

    def _chunk_text(self, text: str, chunk_size: int = 800, overlap_words: int = 120) -> list[str]:
        normalized = re.sub(r"\s+", " ", text).strip()
        if not normalized:
            return []
        words = normalized.split()
        if len(words) <= chunk_size:
            return [normalized]

        chunks: list[str] = []
        start = 0
        while start < len(words):
            end = min(len(words), start + chunk_size)
            chunk = " ".join(words[start:end]).strip()
            if chunk:
                chunks.append(chunk)
            start += max(1, chunk_size - overlap_words)
        return chunks

    async def index_document(self, db: AsyncSession, resource_id: UUID, filename: str | None, content: bytes, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
        resource = await db.get(Resource, resource_id)
        if resource is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Resource not found")

        text = self._extract_text_from_bytes(filename, content)
        text = text.strip()
        if not text:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "No readable text could be extracted from this file")

        chunks = self._chunk_text(text)
        if not chunks:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Document chunking produced no content")

        await db.execute(delete(ResourceChunk).where(ResourceChunk.resource_id == resource_id))

        for index, chunk_text in enumerate(chunks):
            embedding = await self._embed_text(chunk_text)
            db.add(
                ResourceChunk(
                    resource_id=resource_id,
                    chunk_index=index,
                    chunk_text=chunk_text,
                    metadata_={
                        "source_filename": filename or "uploaded-document",
                        "chunk_size": len(chunk_text),
                        "indexed_at": "system",
                        **(metadata or {}),
                    },
                    embedding=embedding,
                )
            )

        await db.commit()
        return {"resource_id": str(resource_id), "chunks_indexed": len(chunks), "status": "indexed"}

    async def rag_query(
        self,
        question: str,
        conversation_id: str | None = None,
        resource_ids: list[str | UUID] | None = None,
        db: AsyncSession | None = None,
        user: User | None = None,
        **_: object,
    ) -> dict:
        if not resource_ids:
            return {
                "answer": "The available documents do not contain enough information to answer that question.",
                "conversation_id": conversation_id,
                "sources": [],
                "retrieval_note": "No resource_ids were supplied; the API intentionally does not search every document in the repository.",
            }

        if db is None or user is None:
            raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Authentication is required for RAG retrieval")

        resolved_ids = []
        for resource_id in resource_ids:
            if resource_id is None:
                continue
            resource_uuid = UUID(str(resource_id))
            await self.ensure_resource_access(db, user, resource_uuid)
            resolved_ids.append(resource_uuid)

        if not resolved_ids:
            raise HTTPException(status.HTTP_403_FORBIDDEN, "No accessible resources were supplied for retrieval")

        chunk_rows = (
            await db.scalars(
                select(ResourceChunk).where(ResourceChunk.resource_id.in_(resolved_ids)).order_by(ResourceChunk.chunk_index)
            )
        ).all()

        if not chunk_rows:
            return {
                "answer": "The available documents do not contain enough information to answer that question.",
                "conversation_id": conversation_id,
                "sources": [],
                "retrieval_note": "The selected documents have not been indexed for retrieval yet.",
            }

        question_vector = await self._embed_text(question)
        resource_lookup = {}
        for resource_id in resolved_ids:
            resource = await db.get(Resource, resource_id)
            if resource is not None:
                resource_lookup[str(resource_id)] = resource.title

        scored_chunks: list[tuple[float, ResourceChunk]] = []
        for chunk in chunk_rows:
            try:
                stored_vector = chunk.embedding or []
                if isinstance(stored_vector, str):
                    stored_vector = json.loads(stored_vector)
            except (json.JSONDecodeError, TypeError):
                stored_vector = []
            if not isinstance(stored_vector, list) or not stored_vector:
                continue
            score = self._cosine_similarity(question_vector, [float(value) for value in stored_vector])
            if math.isfinite(score):
                scored_chunks.append((score, chunk))

        if not scored_chunks:
            return {
                "answer": "The available documents do not contain enough information to answer that question.",
                "conversation_id": conversation_id,
                "sources": [],
                "retrieval_note": "The selected indexed chunks do not contain enough usable vector data for similarity matching.",
            }

        scored_chunks.sort(key=lambda item: item[0], reverse=True)
        context_chunks = [chunk for _, chunk in scored_chunks[:5]]
        context_text = "\n\n".join(
            f"[resource_id={chunk.resource_id}, chunk_index={chunk.chunk_index}] {chunk.chunk_text}"
            for chunk in context_chunks
        )

        prompt = (
            "Answer the user's question only using the retrieved context provided below. "
            "If the retrieved context does not contain enough information, say exactly that the available documents do not contain enough information to answer the question. "
            "Do not invent facts, citations, or resource titles.\n\n"
            f"Retrieved context:\n{context_text}\n\nQuestion: {question}"
        )
        answer = await self._call_gemini(prompt, max_output_tokens=900, temperature=0.2)

        sources = []
        for score, chunk in scored_chunks[:5]:
            sources.append({
                "resource_id": str(chunk.resource_id),
                "title": resource_lookup.get(str(chunk.resource_id), "Unknown resource"),
                "chunk_index": chunk.chunk_index,
                "score": round(score, 4),
            })

        return {
            "answer": answer,
            "conversation_id": conversation_id,
            "sources": sources,
            "retrieval_note": "Context was limited to the requested resource_ids and the most relevant indexed chunks.",
        }

    @staticmethod
    def _cosine_similarity(a: list[float], b: list[float]) -> float:
        if len(a) != len(b):
            length = min(len(a), len(b))
            a = a[:length]
            b = b[:length]
        if not a or not b:
            return 0.0
        dot = sum(x * y for x, y in zip(a, b))
        norm_a = math.sqrt(sum(x * x for x in a))
        norm_b = math.sqrt(sum(x * x for x in b))
        if norm_a == 0 or norm_b == 0:
            return 0.0
        return dot / (norm_a * norm_b)

    async def summarize(
        self,
        filename: str | None,
        content: bytes | str,
        summary_length: str = "medium",
        summary_type: str = "general",
        focus_area: str = "all",
        language: str = "English",
        resource_id: UUID | None = None,
        db: AsyncSession | None = None,
        user: User | None = None,
        **_: object,
    ) -> dict:
        if content is None:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Document content is missing")
        if isinstance(content, (bytes, bytearray)):
            extracted = self._extract_text_from_bytes(filename, bytes(content))
        else:
            extracted = str(content)
        if not extracted.strip():
            extracted = f"Filename: {filename or 'uploaded-document'}\nThe uploaded file is empty or not directly readable as text."
        extracted = extracted[:10000]

        if resource_id is not None and db is not None and user is not None:
            await self.ensure_resource_access(db, user, resource_id)

        prompt = (
            "You are summarizing a research document. Keep the summary clear, factual, and appropriate for a research platform. "
            "Do not invent citations or unavailable details.\n\n"
            f"Document name: {filename or 'uploaded-document'}\n"
            f"Summary length: {summary_length}\n"
            f"Summary type: {summary_type}\n"
            f"Focus area: {focus_area}\n"
            f"Language: {language}\n\n"
            f"Document content:\n{extracted}"
        )
        summary = await self._call_gemini(prompt, max_output_tokens=900, temperature=0.2)
        return {
            "filename": filename or "uploaded-document",
            "summary": summary,
            "summary_length": summary_length,
            "summary_type": summary_type,
            "focus_area": focus_area,
            "language": language,
        }

    async def generate_content(
        self,
        resource_id: UUID | None = None,
        content_type: str = "research_summary",
        topic: str = "",
        tone: str | None = None,
        target_audience: str | None = None,
        language: str = "English",
        instructions: str | None = None,
        db: AsyncSession | None = None,
        user: User | None = None,
        **_: object,
    ) -> dict:
        resource_context = ""
        if resource_id and db is not None and user is not None:
            resource = await self.ensure_resource_access(db, user, resource_id)
            resource_context = (
                f"\nResource context:\n- Title: {resource.title}\n- Type: {resource.resource_type.value}\n- Description: {resource.description or 'No description provided'}\n"
            )

        prompt = (
            "Create polished research communication content for a PolarNexus audience. Keep the output factual and aligned to the topic provided. "
            "Do not invent specific citations, dates, or document sources that are not present in the resource context.\n\n"
            f"Content type: {content_type}\n"
            f"Topic: {topic}\n"
            f"Tone: {tone or 'professional'}\n"
            f"Target audience: {target_audience or 'researchers'}\n"
            f"Language: {language}\n"
            f"Additional instructions: {instructions or 'None'}\n{resource_context}"
        )
        content = await self._call_gemini(prompt, max_output_tokens=1200, temperature=0.5)
        return {
            "content": content,
            "resource_id": str(resource_id) if resource_id else None,
            "content_type": content_type,
            "tone": tone,
            "target_audience": target_audience,
            "language": language,
        }


ai_service = AIService()
